import os
import sys
import xlrd
from docx import Document
from openpyxl.reader.excel import load_workbook


def get_os():
    os_name = sys.platform
    if os_name == "win32" or os_name == "cygwin":
        return "Windows"
    else:
        return "MacOS"


def get_file_names(path):
    file_names = []
    for root, dirs, files in os.walk(path):
        for name in files:
            if not name.startswith('.~'):
                file_names.append(os.path.join(root, name))
    return file_names


def check_file_content(file_path, target_strings):
    target_strings = [x for x in target_strings if x != '']
    content = ''
    file_type = os.path.splitext(file_path)[1]
    if file_type in ['.doc', '.docx']:
        doc = Document(file_path)

        for paragraph in doc.paragraphs:
            content += paragraph.text
    elif file_type == '.xlsx':
        wb = load_workbook(file_path)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    content += str(cell.value)
    elif file_type == '.xls':
        wb = xlrd.open_workbook(file_path)
        sheet = wb.sheet_by_index(0)
        for row in range(sheet.nrows):
            for col in range(sheet.ncols):
                cell_value = sheet.cell_value(row, col)
                if cell_value:
                    content += str(cell_value)
    else:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
    for target_string in target_strings:
        if target_string not in content:
            return False
    return True
